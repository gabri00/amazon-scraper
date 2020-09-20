import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from random import choice

# exel table file
XL_LOG_PATH = 'search_history/search_history.xlsx'

# opens exel file
search_history = pd.read_excel(XL_LOG_PATH)

# coords
x = []
y = []

# colors list
colors = ['red', 'blue', 'orange', 'green', 'cyan']

# font dictionary
font = {'family': 'sans-serif',
		'color':  'darkblue',
		'weight': 'light',
		'size': 12}


def get_row_count() -> int:
	wb = load_workbook(XL_LOG_PATH, enumerate)
	sheet = wb.active
	row_count = sheet.max_row
	wb.close()

	return row_count


def show_menu():
	flag = True

	print('Scegli il prodotto di cui visualizzare il grafico:')

	for it in range(1, rows):
		for i in reversed(range(it-1)):
			if search_history.title[it-1] == search_history.title[i]:
				flag = False

		if not flag:
			break

		print(f'{it}) {search_history.name[it-1]}')


def set_plot():
	plt.style.use('ggplot')
	plt.grid(True)

	# plt.xlabel("Date")
	plt.ylabel("Price")
	plt.title("Price variation of {}".format(search_history.name[prod-1]), fontdict=font)

	for it in range(1, rows):
		if search_history.name[it-1] == search_history.name[prod-1]:
			x.append(search_history.date[it-1])
			y.append(search_history.price[it-1])

	plt.plot(x, y, f'tab:{choice(colors)}')


rows = get_row_count()

while True:
	show_menu()
	print('0) Exit')
	print()

	prod = int(input('Scegli: '))

	if prod == 0:
		break

	set_plot()

	plt.show()